from openpyxl import load_workbook, Workbook
from string import ascii_uppercase as alc
from copy import copy
from pathlib import Path
def create(ws_type='enade',year='2021', courses=[], UFs = [], adm = []):

    my_folder = Path(__file__).parent.resolve()
    enade_wb = load_workbook(filename=my_folder / f'Merged file2.xlsx')
    enade_ws2 = enade_wb.active
    dead_style = copy(enade_ws2['A1'].style)
    idd_wb = load_workbook(filename=my_folder / f'conceito_idd{year}.xlsx')

    workbook = enade_wb
    filename = my_folder / 'novo_conceito_enade.xlsx'
    if ws_type == 'idd':
        filename =my_folder / 'novo_conceito_idd.xlsx'
        workbook =  idd_wb
    enade_ws = workbook.active
    enade_array = []

    new_enade = Workbook()
    new_enade_ws = new_enade.active
    
    header_array = []
    new_enade_ws.row_dimensions[1] = enade_ws.row_dimensions[1]
    for i in alc:
        header_array.append(enade_ws[f'{i}1'].value)
        new_enade_ws.column_dimensions[i] = enade_ws.column_dimensions[i]

    new_enade_ws.append(header_array)

    for i in alc:
        new_enade_ws[f'{i}1'].font = copy(enade_ws[f'{i}1'].font)
        if year != '2018':
            new_enade_ws[f'{i}1'].style = copy(enade_ws[f'{i}1'].style)
        else:
            new_enade_ws[f'{i}1'].style = copy(dead_style)
        new_enade_ws[f'{i}1'].border = copy(enade_ws[f'{i}1'].border)
        new_enade_ws[f'{i}1'].alignment = copy(enade_ws[f'{i}1'].alignment)
    for cell_row in enade_ws.rows:
        try:
            print(cell_row[12].row)
            UFs.index(cell_row[12].value.strip())
            courses.index(cell_row[2].value.strip())
            print(cell_row[12].value, cell_row[2].value)
            enade_array.append(cell_row[0].row)
        except:
            pass
    for data in enade_array:
        print(data)
        cool_array = []
        for chara in alc:
            cool_array.append(enade_ws[f'{chara}{data}'].value)
        new_enade_ws.append(cool_array)
    new_enade.save(filename=filename)
    return filename